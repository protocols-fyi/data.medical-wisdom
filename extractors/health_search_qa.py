from __future__ import annotations

from collections.abc import Iterator
from hashlib import sha256
from itertools import islice
from pathlib import Path

import click
from openpyxl import load_workbook
from tqdm import tqdm

from entities import Pass1Record

DATASET_FILE = Path("datasets.original/health_search_qa/41586_2023_6291_MOESM6_ESM.xlsx")

assert DATASET_FILE.is_file(), f"Missing dataset file: {DATASET_FILE}"


def _questions() -> list[str]:
    workbook = load_workbook(DATASET_FILE, read_only=True, data_only=True)
    questions: list[str] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            value = row[0]
            if value:
                questions.append(str(value).strip())
    return questions


def _record(question: str) -> Pass1Record:
    return Pass1Record.model_validate(
        {
            "id": sha256(question.encode("utf-8")).hexdigest(),
            "question": question,
        }
    )


def extract() -> Iterator[Pass1Record]:
    for question in _questions():
        yield _record(question)


@click.command()
@click.option("--dry-run", is_flag=True, help="Print dataset location and size/count.")
@click.option("--limit", type=int, default=None, help="Only emit the first N parsed records.")
def cli(dry_run: bool, limit: int | None) -> None:
    questions = _questions()
    if dry_run:
        click.echo(f"location: {DATASET_FILE}")
        click.echo(f"size_bytes: {DATASET_FILE.stat().st_size}")
        click.echo(f"questions: {len(questions)}")
        return

    selected = islice(questions, limit) if limit is not None else questions
    total = len(questions) if limit is None else min(limit, len(questions))
    for question in tqdm(selected, total=total, desc="health_search_qa", unit="record"):
        click.echo(_record(question).model_dump_json())


if __name__ == "__main__":
    cli()
